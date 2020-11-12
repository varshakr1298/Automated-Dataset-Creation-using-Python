import pandas as pd
import matplotlib.pyplot as plt
from tabulate import tabulate
from openpyxl import Workbook
from openpyxl import load_workbook
import sys
ch='yes'
filename = "data.xlsx"
wb = load_workbook(filename)
ws = wb.worksheets[0]
global i
i=ws.max_row
i+=1
while(ch=='yes'):
    print("*****************************************************")
    print("1. Play the game")
    print("2. Show the data of the players who played the game")
    print("3. Exit")
    choice=int(input("Please enter your choice from the above menu: "))
    global name
    if choice==1:
        correctAns=0
        wrongAns=0
        name=input("Hello! player kindly enter your name: ")
        print("Question 1: On which date International Day for Biodiversity is observed?")
        print("1) May 15 \n 2) May 20 \n 3) May 22 \n 4) June 5")
        q1=int(input("Enter the correct option: "))
        if(q1==3):
            correctAns+=1
        else:
            wrongAns+=1
        print("Question 2: Which among the following country is considered to have the world’s first sustainable bio-fuels economy?")
        print("1) India \n 2) Brazil \n 3) Australia \n 4) South Africa")
        q2=int(input("Enter the correct option: "))
        if(q2==2):
            correctAns+=1
        else:
            wrongAns+=1
        print("Question 3: Largest share in global mangrove areas are found in?")
        print("1) India \n 2) Brazil \n 3) Indonesia \n 4) South Africa")
        q3=int(input("Enter the correct option: "))
        if(q3==3):
            correctAns+=1
        else:
            wrongAns+=1
        category=["Correct Answers","Wrong Answers"]
        colors=["m","b"]
        explode=(0.1,0)
        numberOfQuestions=[correctAns,wrongAns]
        plt.pie(numberOfQuestions,explode=explode,labels=category,colors=colors,autopct="%1.1f%%")
        plt.title("Result Analysis")
        plt.show()
        print("Your total score: "+str(correctAns)+"/3")
        wcell1=ws.cell(i,1)
        wcell1.value=name
        wcell2=ws.cell(i,2)
        wcell2.value=correctAns
        wcell3=ws.cell(i,3)
        wcell3.value=wrongAns
        wcell4=ws.cell(i,4)
        wcell4.value=3
        i+=1
        wb.save(filename)
    elif choice==2:
        headers=['Name','Number of Correct Answers','Number of Wrong Answers','Total Questions']
        df=pd.read_excel(r'data.xlsx',names=headers,header=None)
        print(tabulate(df,headers="keys",tablefmt="grid"))
    elif choice==3:
        sys.exit(0)
    
    print("**********************************************************")
    ch=input("Do you wish to continue(yes/no): ").lower()
    if(ch!='yes'):
        break
